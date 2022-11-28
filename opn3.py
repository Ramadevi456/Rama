from openpyxl import Workbook
#import datatime
wb=Workbook()
ws=wb.active
ws['A1'] = "Invoice"
ws['B1'] = "LinePrice"
ws['C1'] = "Discription"
ws['D1'] = "Quantity"
ws['E1'] = "InvoiceData"
ws['F1'] = "UnitPrice"
ws['G1'] = "CustomerID"
ws['H1'] = "Country"
ws.append([11,12,13,33,55,76,45,56])
ws.append([31,32,33,13,99,67,55,66])
ws.append([41,42,43,45,44,47,49,99])

ws.append([51,52,53,77,56,44,78,88])

import datetime
#ws['A2'] = datetime.datetime.now()
wb.save("sample.xlsx")

